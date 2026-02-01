#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Вариант 2 (агрегаты):

- moves_by_date.json может хранить только ~90 дней.
- Для корректного "нарастающего" ведём monthly_aggregates.json:
  { "YYYY-MM": stats_like_table, ... }

Отчёт (вся таблица) считается ТОЛЬКО по завершённым карточкам (ГПЗУ/ОТКАЗ).
РПГУ всегда 0. ПРИОСТАНОВЛЕННЫХ всегда 0.

Telegram: сообщение "Отчёт готов за <Месяц> <Год>." и Excel (.xlsx) с 3 вкладками:
- Ежемесячный
- Нарастающим
- Детализация (только за месяц)
"""

import os
import json
import time
import threading
from pathlib import Path
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


NOVOKUZNETSK_TZ = ZoneInfo("Asia/Novokuznetsk")

# Финальные колонки (завершение)
POSITIVE_TO_COLUMN_IDS = {5474978}  # ГПЗУ
REFUSAL_TO_COLUMN_IDS = {5474969}   # ОТКАЗ

PERSON_PHYS = "Физическое лицо"
PERSON_JUR = "Юридическое лицо"

SUBM_EPGU = "ЕПГУ"
SUBM_MFC = "МФЦ"
SUBM_PERSONAL = "Личный приём"
SUBM_RPGU = "РПГУ"  # всегда 0

RU_MONTHS = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}


# ====== Даты/календарь ======
def date_to_str(d: date) -> str:
    return d.strftime("%Y-%m-%d")

def month_key(d: date) -> str:
    return d.strftime("%Y-%m")

def is_weekend(d: date) -> bool:
    return d.weekday() >= 5

def is_workday(d: date, holidays: set[str]) -> bool:
    return (not is_weekend(d)) and (date_to_str(d) not in holidays)

def first_workday_of_month(year: int, month: int, holidays: set[str]) -> date:
    d = date(year, month, 1)
    while not is_workday(d, holidays):
        d += timedelta(days=1)
    return d

def prev_month_range(today: date) -> tuple[date, date]:
    first_this = date(today.year, today.month, 1)
    last_prev = first_this - timedelta(days=1)
    first_prev = date(last_prev.year, last_prev.month, 1)
    return first_prev, last_prev


# ====== Paths ======
def aggregates_path_for(data_file_path: str) -> Path:
    return Path(data_file_path).with_name("monthly_aggregates.json")

def state_path_for(data_file_path: str) -> Path:
    return Path(data_file_path).with_name("monthly_report_state.json")


# ====== JSON load/save ======
def _safe_load_json(path: Path, default):
    try:
        if not path.exists():
            return default
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default

def _safe_save_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


# ====== State (чтобы не слать повторно) ======
def load_state(state_path: Path) -> dict:
    return _safe_load_json(state_path, {})

def save_state(state_path: Path, state: dict) -> None:
    _safe_save_json(state_path, state)


# ====== Агрегаты ======
def load_aggregates(aggr_path: Path) -> dict:
    return _safe_load_json(aggr_path, {})

def save_aggregates(aggr_path: Path, aggr: dict) -> None:
    _safe_save_json(aggr_path, aggr)

def _zero_stats_like_table() -> dict:
    return {
        "accepted_total": {"phys": 0, "jur": 0},
        "accepted_by_method": {
            SUBM_PERSONAL: {"phys": 0, "jur": 0},
            SUBM_MFC:      {"phys": 0, "jur": 0},
            SUBM_EPGU:     {"phys": 0, "jur": 0},
            SUBM_RPGU:     {"phys": 0, "jur": 0},  # всегда 0
        },
        "services_total": {"phys": 0, "jur": 0},
        "positive": {"phys": 0, "jur": 0},
        "suspended": {"phys": 0, "jur": 0},  # всегда 0
        "refusals": {"phys": 0, "jur": 0},
    }

def _add_stats(dst: dict, src: dict) -> dict:
    # Суммируем только числовые поля структуры
    for k in ("accepted_total", "services_total", "positive", "suspended", "refusals"):
        dst[k]["phys"] += int(src.get(k, {}).get("phys", 0))
        dst[k]["jur"]  += int(src.get(k, {}).get("jur", 0))

    for m in (SUBM_PERSONAL, SUBM_MFC, SUBM_EPGU, SUBM_RPGU):
        dst["accepted_by_method"][m]["phys"] += int(src.get("accepted_by_method", {}).get(m, {}).get("phys", 0))
        dst["accepted_by_method"][m]["jur"]  += int(src.get("accepted_by_method", {}).get(m, {}).get("jur", 0))
    return dst

def sum_aggregates_ytd(aggr: dict, year: int, end_month: int) -> dict:
    """
    Сумма агрегатов с января по end_month включительно.
    """
    total = _zero_stats_like_table()
    for m in range(1, end_month + 1):
        key = f"{year}-{m:02d}"
        if key in aggr:
            total = _add_stats(total, aggr[key])
    # Жёстко фиксируем требования
    total["accepted_by_method"][SUBM_RPGU] = {"phys": 0, "jur": 0}
    total["suspended"] = {"phys": 0, "jur": 0}
    return total


# ====== Чтение moves по диапазону дат ======
def _iter_moves_in_range(store: dict, start: date, end: date):
    d = start
    while d <= end:
        day = store.get(date_to_str(d), {})
        moves = day.get("moves", [])
        if isinstance(moves, list):
            for m in moves:
                if isinstance(m, dict):
                    yield m
        d += timedelta(days=1)


def _normalize_person_type(raw: str) -> str:
    if raw and "Физ" in raw:
        return PERSON_PHYS
    return PERSON_JUR


def _normalize_submission_method(raw: str) -> str:
    """
    Способ подачи для таблицы.
    РПГУ принудительно считаем 0 (по требованию): если встретится, будем игнорировать в подсчёте.
    """
    if not raw:
        return SUBM_EPGU
    if "ЕПГУ" in raw:
        return SUBM_EPGU
    if "МФЦ" in raw:
        return SUBM_MFC
    if "Лич" in raw:
        return SUBM_PERSONAL
    if "РПГУ" in raw:
        return SUBM_RPGU
    return SUBM_EPGU


def _collect_completed_last_event(store: dict, start: date, end: date) -> dict:
    """
    Завершённые карточки за период. Берём последнее завершение в периоде.
    card_id -> {person_type, submission_method, result, timestamp, title}
    """
    completed: dict[str, dict] = {}

    for m in _iter_moves_in_range(store, start, end):
        card_id = m.get("card_id")
        if not card_id:
            continue

        to_col = m.get("to_column_id")
        if to_col not in (POSITIVE_TO_COLUMN_IDS | REFUSAL_TO_COLUMN_IDS):
            continue

        ts = m.get("timestamp", "") or ""
        person = _normalize_person_type(m.get("person_type", ""))
        method = _normalize_submission_method(m.get("submission_method", ""))

        result = "ГПЗУ" if to_col in POSITIVE_TO_COLUMN_IDS else "ОТКАЗ"

        rec = {
            "card_id": str(card_id),
            "title": (m.get("title", "") or "").replace("\n", " ").strip(),
            "person_type": person,
            "submission_method": method,
            "result": result,
            "timestamp": ts,
        }

        prev = completed.get(rec["card_id"])
        if prev is None or rec["timestamp"] > prev.get("timestamp", ""):
            completed[rec["card_id"]] = rec

    return completed


# ====== Расчёт по завершённым (форма прежняя, без линий-разделителей) ======
def compute_month_like_stats(store: dict, start: date, end: date) -> dict:
    """
    Вся таблица считается ТОЛЬКО по завершённым карточкам за период.
    """
    completed = _collect_completed_last_event(store, start, end)

    accepted_total = {"phys": 0, "jur": 0}
    accepted_by_method = {
        SUBM_PERSONAL: {"phys": 0, "jur": 0},
        SUBM_MFC:      {"phys": 0, "jur": 0},
        SUBM_EPGU:     {"phys": 0, "jur": 0},
        SUBM_RPGU:     {"phys": 0, "jur": 0},  # всегда 0
    }

    positive = {"phys": 0, "jur": 0}
    refusals = {"phys": 0, "jur": 0}

    for rec in completed.values():
        who = "phys" if rec["person_type"] == PERSON_PHYS else "jur"

        accepted_total[who] += 1

        method = rec["submission_method"]
        if method != SUBM_RPGU:  # РПГУ всегда 0
            if method not in accepted_by_method:
                method = SUBM_EPGU
            accepted_by_method[method][who] += 1

        if rec["result"] == "ГПЗУ":
            positive[who] += 1
        else:
            refusals[who] += 1

    services_total = {
        "phys": positive["phys"] + refusals["phys"],
        "jur":  positive["jur"] + refusals["jur"],
    }

    return {
        "accepted_total": accepted_total,
        "accepted_by_method": accepted_by_method,
        "services_total": services_total,
        "positive": positive,
        "suspended": {"phys": 0, "jur": 0},
        "refusals": refusals,
    }


def _pct(part: int, total: int) -> str:
    if total <= 0:
        return "0%"
    return f"{round((part / total) * 100):d}%"


def render_month_table(title: str, stats: dict) -> str:
    """
    Текстовый рендер формы (без строк '-----', строго подряд).
    Проценты считаются отдельно для физ. и юр. относительно их итогов.
    """
    acc_phys = stats["accepted_total"]["phys"]
    acc_jur  = stats["accepted_total"]["jur"]

    def row(label, phys, jur):
        return f"{label:<34} {str(phys):>6} {str(jur):>6}"

    lines = []
    lines.append(f"{title}")
    lines.append("```")
    lines.append(f"{'':<34} {'физ.':>6} {'юр.':>6}")

    def pct_phys(part: int) -> str:
        return _pct(part, acc_phys)

    def pct_jur(part: int) -> str:
        return _pct(part, acc_jur)

    lines.append(row("Общее количество принятых заявлений", acc_phys, acc_jur))

    lp = stats["accepted_by_method"][SUBM_PERSONAL]
    lines.append(row("лично в ОМС", lp["phys"], lp["jur"]))
    lines.append(row("% от общего числа", pct_phys(lp["phys"]), pct_jur(lp["jur"])))

    mfc = stats["accepted_by_method"][SUBM_MFC]
    lines.append(row("через МФЦ", mfc["phys"], mfc["jur"]))
    lines.append(row("% от общего числа", pct_phys(mfc["phys"]), pct_jur(mfc["jur"])))

    epgu = stats["accepted_by_method"][SUBM_EPGU]
    lines.append(row("через ЕПГУ", epgu["phys"], epgu["jur"]))
    lines.append(row("% от общего числа", pct_phys(epgu["phys"]), pct_jur(epgu["jur"])))

    rpgu = stats["accepted_by_method"][SUBM_RPGU]
    lines.append(row("через РПГУ", rpgu["phys"], rpgu["jur"]))
    # РПГУ по вашей логике всегда 0
    lines.append(row("% от общего числа", "0%", "0%"))

    st = stats["services_total"]
    pos = stats["positive"]
    sus = stats["suspended"]
    ref = stats["refusals"]

    lines.append(row("Общее количество оказанных услуг", st["phys"], st["jur"]))
    lines.append(row("ПОЛОЖИТЕЛЬНЫХ", pos["phys"], pos["jur"]))
    lines.append(row("ПРИОСТАНОВЛЕННЫХ", sus["phys"], sus["jur"]))
    lines.append(row("ОТКАЗЫ", ref["phys"], ref["jur"]))

    lines.append("```")
    return "\n".join(lines)



# ====== Excel: 3 вкладки ======
def _autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)


def _write_form_sheet(ws, title: str, stats: dict):
    ws.title = ws.title[:31]
    ws["A1"] = title

    ws.append(["", "физ.", "юр."])

    acc_phys = stats["accepted_total"]["phys"]
    acc_jur = stats["accepted_total"]["jur"]

    def pct_phys(part: int) -> str:
        if acc_phys <= 0:
            return "0%"
        return f"{round((part / acc_phys) * 100):d}%"

    def pct_jur(part: int) -> str:
        if acc_jur <= 0:
            return "0%"
        return f"{round((part / acc_jur) * 100):d}%"

    ws.append(["Общее количество принятых заявлений", acc_phys, acc_jur])

    lp = stats["accepted_by_method"][SUBM_PERSONAL]
    ws.append(["лично в ОМС", lp["phys"], lp["jur"]])
    ws.append(["% от общего числа", pct_phys(lp["phys"]), pct_jur(lp["jur"])])

    mfc = stats["accepted_by_method"][SUBM_MFC]
    ws.append(["через МФЦ", mfc["phys"], mfc["jur"]])
    ws.append(["% от общего числа", pct_phys(mfc["phys"]), pct_jur(mfc["jur"])])

    epgu = stats["accepted_by_method"][SUBM_EPGU]
    ws.append(["через ЕПГУ", epgu["phys"], epgu["jur"]])
    ws.append(["% от общего числа", pct_phys(epgu["phys"]), pct_jur(epgu["jur"])])

    rpgu = stats["accepted_by_method"][SUBM_RPGU]
    ws.append(["через РПГУ", rpgu["phys"], rpgu["jur"]])
    ws.append(["% от общего числа", "0%", "0%"])

    st = stats["services_total"]
    pos = stats["positive"]
    sus = stats["suspended"]
    ref = stats["refusals"]

    ws.append(["Общее количество оказанных услуг", st["phys"], st["jur"]])
    ws.append(["ПОЛОЖИТЕЛЬНЫХ", pos["phys"], pos["jur"]])
    ws.append(["ПРИОСТАНОВЛЕННЫХ", sus["phys"], sus["jur"]])
    ws.append(["ОТКАЗЫ", ref["phys"], ref["jur"]])

    _autosize_columns(ws)



def build_month_excel_report(store: dict, month_start: date, month_end: date, aggregates_path: Path) -> tuple[str, bytes]:
    """
    Excel:
      - Ежемесячный: считаем из store (по завершённым)
      - Нарастающим: суммируем monthly_aggregates.json (включая текущий месяц, который сейчас обновим)
      - Детализация: только месяц (из store)
    """
    key = month_key(month_end)

    # 1) Считаем месяц из store
    month_stats = compute_month_like_stats(store, month_start, month_end)

    # 2) Обновляем агрегаты для этого месяца (чтобы нарастающий был корректным)
    aggr = load_aggregates(aggregates_path)
    aggr[key] = month_stats
    save_aggregates(aggregates_path, aggr)

    # 3) Считаем YTD из агрегатов
    ytd_stats = sum_aggregates_ytd(aggr, month_end.year, month_end.month)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Ежемесячный"
    ws2 = wb.create_sheet("Нарастающим")
    ws3 = wb.create_sheet("Детализация")

    month_title = f"Ежемесячный отчёт за {RU_MONTHS[month_end.month]} {month_end.year}"
    ytd_title = f"Нарастающим итогом: {RU_MONTHS[1]}–{RU_MONTHS[month_end.month]} {month_end.year}"

    _write_form_sheet(ws1, month_title, month_stats)
    _write_form_sheet(ws2, ytd_title, ytd_stats)

    # Детализация (только месяц)
    completed = _collect_completed_last_event(store, month_start, month_end)
    rows = sorted(completed.values(), key=lambda x: x.get("timestamp", ""))

    ws3.append(["timestamp", "статус", "тип", "подача", "card_id", "заявитель"])
    for r in rows:
        typ = "физ" if r["person_type"] == PERSON_PHYS else "юр"
        ws3.append([
            (r.get("timestamp") or ""),
            r.get("result") or "",
            typ,
            r.get("submission_method") or "",
            r.get("card_id") or "",
            r.get("title") or "",
        ])
    _autosize_columns(ws3)

    filename = f"report_{month_end.year}-{month_end.month:02d}.xlsx"

    import io
    bio = io.BytesIO()
    wb.save(bio)
    return filename, bio.getvalue()


# ====== Telegram: отправка файла ======
def send_telegram_document(file_bytes: bytes, filename: str) -> bool:
    token = os.getenv("TELEGRAM_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        print("[TELEGRAM] ⚠️ Telegram не настроен (TELEGRAM_TOKEN/TELEGRAM_CHAT_ID)")
        return False

    url = f"https://api.telegram.org/bot{token}/sendDocument"

    try:
        files = {"document": (filename, file_bytes,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"chat_id": chat_id}
        resp = requests.post(url, data=data, files=files, timeout=60)
        if resp.status_code == 200:
            return True
        print(f"[TELEGRAM] ❌ sendDocument статус={resp.status_code}, ответ={resp.text[:300]}")
        return False
    except Exception as e:
        print(f"[TELEGRAM] ❌ Ошибка sendDocument: {e}")
        return False


# ====== Авто-рассылка 1-й рабочий день ======
def start_monthly_reports_thread(
    *,
    load_store_func,
    send_telegram_func,
    data_file_path: str,
    holidays: set[str],
    tz: ZoneInfo = NOVOKUZNETSK_TZ,
    send_time_hm: tuple[int, int] = (8, 35),
) -> threading.Thread:
    st_path = state_path_for(data_file_path)
    aggr_path = aggregates_path_for(data_file_path)

    def loop():
        print("[MONTHLY] 🤖 Вариант 2: агрегаты monthly_aggregates.json для нарастающего.")
        print(f"[MONTHLY] AGGREGATES={aggr_path}")
        print(f"[MONTHLY] STATE={st_path}")

        while True:
            try:
                now = datetime.now(tz)
                hh, mm = send_time_hm
                target = now.replace(hour=hh, minute=mm, second=0, microsecond=0)
                if now >= target:
                    target += timedelta(days=1)
                time.sleep((target - now).total_seconds())

                today = datetime.now(tz).date()
                fwd = first_workday_of_month(today.year, today.month, holidays)
                if today != fwd:
                    continue

                pm_start, pm_end = prev_month_range(today)
                pm_key = month_key(pm_end)

                state = load_state(st_path)
                if state.get("last_sent") == pm_key:
                    continue

                store = load_store_func()

                filename, xlsx_bytes = build_month_excel_report(store, pm_start, pm_end, aggr_path)

                # Сообщение + файл
                send_telegram_func(f"Отчёт готов за {RU_MONTHS[pm_end.month]} {pm_end.year}.")
                ok = send_telegram_document(xlsx_bytes, filename)
                if not ok:
                    print("[MONTHLY] ❌ Не удалось отправить Excel.")
                    continue

                state["last_sent"] = pm_key
                save_state(st_path, state)
                print(f"[MONTHLY] ✅ Отправлено за {pm_key}")

            except Exception as e:
                print(f"[MONTHLY] ❌ Ошибка: {e}")
                time.sleep(60)

    t = threading.Thread(target=loop, daemon=True)
    t.start()
    return t
